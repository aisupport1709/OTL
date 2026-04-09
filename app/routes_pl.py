import os
import io
from flask import Blueprint, render_template, request, jsonify, current_app, send_file
from werkzeug.utils import secure_filename
from app import db
from app.models.pl_entry import PLEntry
from app.models.pl_sdck import PLSDCK
from app.models.account_mapping import AccountMapping
from app.services.pl_import import (
    import_pl_file, import_pl_google_sheet, calculate_pl, get_available_years
)
from app.routes_auth import login_required, app_access_required

pl_bp = Blueprint('pl', __name__, url_prefix='/pl')
pl_api_bp = Blueprint('pl_api', __name__, url_prefix='/pl/api')

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ─── Pages ───────────────────────────────────────────────────────────

@pl_bp.route('/')
@login_required
@app_access_required('pl')
def report_page():
    return render_template('pl/report.html')


@pl_bp.route('/import')
@login_required
@app_access_required('pl')
def import_page():
    return render_template('pl/import.html')


# ─── API: Upload & Import ────────────────────────────────────────────

@pl_api_bp.route('/upload/pl-file', methods=['POST'])
@login_required
@app_access_required('pl')
def upload_pl_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'Only .xlsx and .xls files are allowed'}), 400

    filename = secure_filename(file.filename)
    upload_folder = current_app.config['UPLOAD_FOLDER']
    os.makedirs(upload_folder, exist_ok=True)
    filepath = os.path.join(upload_folder, filename)
    file.save(filepath)

    try:
        success, errors_count, error_details = import_pl_file(filepath, filename)
        return jsonify({
            'message': f'Import completed: {success} records imported, {errors_count} errors',
            'success_count': success,
            'error_count': errors_count,
            'errors': error_details[:20],
        })
    except Exception as e:
        return jsonify({'error': f'Import failed: {str(e)}'}), 500
    finally:
        if os.path.exists(filepath):
            os.remove(filepath)


@pl_api_bp.route('/upload/pl-google', methods=['POST'])
@login_required
@app_access_required('pl')
def upload_pl_google():
    data = request.get_json()
    url = (data or {}).get('url', '').strip()
    month = (data or {}).get('month')
    year = (data or {}).get('year')

    if not url:
        return jsonify({'error': 'No Google Sheets URL provided'}), 400

    if 'docs.google.com/spreadsheets' not in url:
        return jsonify({'error': 'Invalid Google Sheets URL'}), 400

    try:
        success, errors_count, error_details = import_pl_google_sheet(url, month=month, year=year)
        return jsonify({
            'message': f'Import completed: {success} records imported, {errors_count} errors',
            'success_count': success,
            'error_count': errors_count,
            'errors': error_details[:20],
        })
    except Exception as e:
        error_msg = str(e)
        # Improve error message for common issues
        if '404' in error_msg:
            error_msg = 'Sheet not found. Make sure: (1) URL is correct, (2) Sheet is publicly shared (Share → Anyone with the link), (3) Sheet exists and is not deleted'
        elif 'HTTP Error' in error_msg:
            error_msg = 'Cannot access Google Sheet. Ensure it is publicly shared (Share → Anyone with the link can view)'

        return jsonify({'error': f'Import failed: {error_msg}'}), 500


# ─── API: P&L Report ────────────────────────────────────────────────

@pl_api_bp.route('/report', methods=['GET'])
@login_required
@app_access_required('pl')
def get_pl_report():
    year = request.args.get('year', type=int)

    if not year:
        return jsonify({'error': 'Year parameter required'}), 400

    try:
        pl_data = calculate_pl(year)

        # Keys that are COGS sub-components, not standalone P&L line items
        COGS_SUB_KEYS = {'cogs_opening', 'cogs_production', 'cogs_ending'}
        COGS_SENTINEL = {
            'cogs_opening':    '__opening__',
            'cogs_production': '__production__',
            'cogs_ending':     '__ending__',
        }
        COGS_LABELS = {
            'cogs_opening':    'Giá vốn hàng bán đầu kỳ',
            'cogs_production': 'Giá vốn hàng bán phát sinh trong kỳ',
            'cogs_ending':     'Giá vốn hàng bán số dư cuối kỳ',
        }

        annual_totals = {
            'Doanh thu thuần': 0, 'Giá vốn hàng bán': 0, 'Lợi nhuận gộp': 0,
            'Chi phí QLDN': 0, 'Lợi nhuận thuần từ HĐKD': 0, 'Doanh thu HĐTC': 0,
            'Chi phí tài chính': 0, 'Thu nhập khác': 0, 'Chi phí khác': 0,
            'Lợi nhuận trước thuế': 0, 'Thuế TNDN': 0,
        }

        # cogs_groups: sentinel_key → {name, monthly: {month: val}, annual, sub: {code: {name, monthly, annual}}}
        cogs_groups = {
            sentinel: {'name': COGS_LABELS[key], 'monthly': {}, 'annual': 0, 'sub': {}}
            for key, sentinel in COGS_SENTINEL.items()
        }

        cleaned_monthly = {}
        sub_accounts_by_line = {}

        for month, month_data in pl_data.items():
            cleaned_monthly[month] = {}
            sub_accts = month_data.pop('__sub_accounts__', {})

            for key, value in month_data.items():
                if key in COGS_SUB_KEYS:
                    # Track COGS sub-components separately
                    sentinel = COGS_SENTINEL[key]
                    cogs_groups[sentinel]['monthly'][month] = value
                    cogs_groups[sentinel]['annual'] += value
                else:
                    cleaned_monthly[month][key] = value
                    if key in annual_totals:
                        annual_totals[key] += value

            # Aggregate sub_accounts
            for line_item, codes_dict in sub_accts.items():
                if line_item in COGS_SUB_KEYS:
                    # Place into the correct cogs_group sub
                    sentinel = COGS_SENTINEL[line_item]
                    for code, details in codes_dict.items():
                        grp_sub = cogs_groups[sentinel]['sub']
                        if code not in grp_sub:
                            grp_sub[code] = {'name': details['name'], 'monthly': {}, 'annual': 0}
                        grp_sub[code]['monthly'][month] = details['value']
                        grp_sub[code]['annual'] += details['value']
                else:
                    if line_item not in sub_accounts_by_line:
                        sub_accounts_by_line[line_item] = {}
                    for code, details in codes_dict.items():
                        if code not in sub_accounts_by_line[line_item]:
                            sub_accounts_by_line[line_item][code] = {
                                'name': details['name'], 'monthly': {}, 'annual': 0
                            }
                        sub_accounts_by_line[line_item][code]['monthly'][month] = details['value']
                        sub_accounts_by_line[line_item][code]['annual'] += details['value']

        # Attach COGS groups as sentinel-keyed entries under 'Giá vốn hàng bán'
        sub_accounts_by_line['Giá vốn hàng bán'] = cogs_groups

        return jsonify({
            'year': year,
            'monthly': cleaned_monthly,
            'annual': annual_totals,
            'sub_accounts': sub_accounts_by_line,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to calculate P&L: {str(e)}'}), 500


@pl_api_bp.route('/available-years', methods=['GET'])
@login_required
@app_access_required('pl')
def get_available_years_api():
    try:
        years = get_available_years()
        return jsonify({'years': years})
    except Exception as e:
        return jsonify({'error': f'Failed to fetch available years: {str(e)}'}), 500


@pl_api_bp.route('/export', methods=['GET'])
@login_required
@app_access_required('pl')
def export_pl_report():
    """Export P&L report to Excel file."""
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    year = request.args.get('year', type=int)

    if not year:
        return jsonify({'error': 'Year parameter required'}), 400

    try:
        pl_data = calculate_pl(year)

        COGS_SUB_KEYS = {'cogs_opening', 'cogs_production', 'cogs_ending'}
        COGS_SENTINEL = {
            'cogs_opening':    '__opening__',
            'cogs_production': '__production__',
            'cogs_ending':     '__ending__',
        }
        COGS_GROUP_LABELS_EN = {
            '__opening__':    'Opening balance',
            '__production__': 'Purchases',
            '__ending__':     'Closing balance',
        }
        COGS_LABELS = {
            'cogs_opening':    'Giá vốn hàng bán đầu kỳ',
            'cogs_production': 'Giá vốn hàng bán phát sinh trong kỳ',
            'cogs_ending':     'Giá vốn hàng bán số dư cuối kỳ',
        }

        annual_totals = {
            'Doanh thu thuần': 0, 'Giá vốn hàng bán': 0, 'Lợi nhuận gộp': 0,
            'Chi phí QLDN': 0, 'Lợi nhuận thuần từ HĐKD': 0, 'Doanh thu HĐTC': 0,
            'Chi phí tài chính': 0, 'Thu nhập khác': 0, 'Chi phí khác': 0,
            'Lợi nhuận trước thuế': 0, 'Thuế TNDN': 0,
        }
        cogs_groups = {
            sentinel: {'name': COGS_LABELS[key], 'monthly': {}, 'annual': 0, 'sub': {}}
            for key, sentinel in COGS_SENTINEL.items()
        }
        cleaned_monthly = {}
        sub_accounts_by_line = {}

        for month, month_data in pl_data.items():
            cleaned_monthly[month] = {}
            sub_accts = month_data.pop('__sub_accounts__', {})

            for key, value in month_data.items():
                if key in COGS_SUB_KEYS:
                    sentinel = COGS_SENTINEL[key]
                    cogs_groups[sentinel]['monthly'][month] = value
                    cogs_groups[sentinel]['annual'] += value
                else:
                    cleaned_monthly[month][key] = value
                    if key in annual_totals:
                        annual_totals[key] += value

            for line_item, codes_dict in sub_accts.items():
                if line_item in COGS_SUB_KEYS:
                    sentinel = COGS_SENTINEL[line_item]
                    for code, details in codes_dict.items():
                        grp_sub = cogs_groups[sentinel]['sub']
                        if code not in grp_sub:
                            grp_sub[code] = {'name': details['name'], 'monthly': {}, 'annual': 0}
                        grp_sub[code]['monthly'][month] = details['value']
                        grp_sub[code]['annual'] += details['value']
                else:
                    if line_item not in sub_accounts_by_line:
                        sub_accounts_by_line[line_item] = {}
                    for code, details in codes_dict.items():
                        if code not in sub_accounts_by_line[line_item]:
                            sub_accounts_by_line[line_item][code] = {
                                'name': details['name'], 'monthly': {}, 'annual': 0
                            }
                        sub_accounts_by_line[line_item][code]['monthly'][month] = details['value']
                        sub_accounts_by_line[line_item][code]['annual'] += details['value']

        sub_accounts_by_line['Giá vốn hàng bán'] = cogs_groups

        # Load account mappings for EN name column
        all_mappings = {m.local_code: m.hq_code for m in AccountMapping.query.all()}

        # Build Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f'P&L {year}'

        # Define styles
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        main_row_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        main_row_font = Font(bold=True, size=10)
        subtotal_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        subtotal_font = Font(bold=True, size=10)
        sub_font = Font(size=9)
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        MONTH_NAMES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                       'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

        SUBTOTAL_ITEMS = {'Lợi nhuận gộp', 'Lợi nhuận thuần từ HĐKD', 'Lợi nhuận trước thuế'}
        # (name_key, en_label, acc_code, is_computed)
        LINE_ITEMS_EXPORT = [
            ('Doanh thu thuần',          'REVENUE',                                                    None,   False),
            ('Giá vốn hàng bán',         'COST OF GOODS SOLD',                                         None,   False),
            ('Lợi nhuận gộp',            'Gross profit',                                               None,   False),
            ('Chi phí QLDN',             'ADMINISTRATION EXPENSES (WITH RED INVOICES)',                 None,   False),
            ('Lợi nhuận thuần từ HĐKD', 'OPERATING PROFIT FROM BUSINESS ACTIVITIES (TAX PURPOSE)',    None,   False),
            ('Doanh thu HĐTC',           'FINANCIAL INCOME',                                           None,   False),
            ('Chi phí tài chính',        'FINANCIAL EXPENSES',                                         None,   False),
            ('Thu nhập khác',            'OTHER INCOMES',                                              None,   False),
            ('Chi phí khác',             'OTHER EXPENSES (Without RED INVOICES)',                      None,   False),
            ('Lợi nhuận trước thuế',     'PROFIT/(LOSS) BEFORE TAXATION',                              None,   False),
            ('Thuế TNDN',                'INCOME TAX',                                                 '8211', False),
            ('__profit_after_tax__',     'PROFIT AFTER TAXATION',                                      None,   True),
        ]

        # Column layout: A=Description, B=Acc No, C=Desc EN, D..O=months, P=Annual FS
        # months start at col 4 (D), annual at col 16 (P)
        MONTH_START_COL = 4
        ANNUAL_COL = MONTH_START_COL + 12  # 16

        # Header row
        ws['A1'] = 'Description'
        ws['B1'] = 'Acc No'
        ws['C1'] = 'Account name (EN)'
        for col_letter_hdr, hdr_val in [('A', 'Description'), ('B', 'Acc No'), ('C', 'Desc EN')]:
            c = ws[f'{col_letter_hdr}1']
            c.value = hdr_val
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
            c.border = thin_border

        for i, month_name in enumerate(MONTH_NAMES):
            col_letter = get_column_letter(MONTH_START_COL + i)
            cell = ws[f'{col_letter}1']
            cell.value = month_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Annual column
        annual_col = get_column_letter(ANNUAL_COL)
        ws[f'{annual_col}1'] = 'Annual FS'
        ws[f'{annual_col}1'].font = header_font
        ws[f'{annual_col}1'].fill = header_fill
        ws[f'{annual_col}1'].alignment = center_align
        ws[f'{annual_col}1'].border = thin_border

        # Data rows
        row = 2
        for (line_item, en_label, acc_code, is_computed) in LINE_ITEMS_EXPORT:
            is_subtotal = (line_item in SUBTOTAL_ITEMS) or is_computed
            row_font = subtotal_font if is_subtotal else main_row_font
            row_fill = subtotal_fill if is_subtotal else main_row_fill

            # EN name from mapping if acc_code known
            en_name = all_mappings.get(acc_code, '') if acc_code else ''

            ws[f'A{row}'] = en_label
            ws[f'B{row}'] = acc_code or ''
            ws[f'C{row}'] = en_name
            for col_key in ('A', 'B', 'C'):
                ws[f'{col_key}{row}'].font = row_font
                ws[f'{col_key}{row}'].fill = row_fill
                ws[f'{col_key}{row}'].border = thin_border

            for i in range(12):
                col_letter = get_column_letter(MONTH_START_COL + i)
                cell = ws[f'{col_letter}{row}']
                if is_computed:
                    value = (cleaned_monthly.get(i + 1, {}).get('Lợi nhuận trước thuế', 0)
                             - cleaned_monthly.get(i + 1, {}).get('Thuế TNDN', 0))
                else:
                    value = cleaned_monthly.get(i + 1, {}).get(line_item, 0)
                cell.value = value if value != 0 else None
                cell.font = row_font
                cell.fill = row_fill
                cell.alignment = right_align
                cell.number_format = '#,##0;(#,##0)'
                cell.border = thin_border

            annual_cell = ws[f'{annual_col}{row}']
            if is_computed:
                annual_value = (annual_totals.get('Lợi nhuận trước thuế', 0)
                                - annual_totals.get('Thuế TNDN', 0))
            else:
                annual_value = annual_totals.get(line_item, 0)
            annual_cell.value = annual_value if annual_value != 0 else None
            annual_cell.font = row_font
            annual_cell.fill = row_fill
            annual_cell.alignment = right_align
            annual_cell.number_format = '#,##0;(#,##0)'
            annual_cell.border = thin_border

            row += 1

            if is_computed:
                continue  # no sub-rows for computed items

            # Sub-account rows
            subs = sub_accounts_by_line.get(line_item, {})

            if line_item == 'Giá vốn hàng bán':
                cogs_sub_header_fill = PatternFill(start_color='EBF3FB', end_color='EBF3FB', fill_type='solid')
                cogs_sub_header_font = Font(bold=True, italic=True, size=9)
                cogs_leaf_font = Font(size=9)

                COGS_ORDER = ['__opening__', '__production__', '__ending__']
                for sentinel in COGS_ORDER:
                    group = subs.get(sentinel)
                    if not group:
                        continue

                    ws[f'A{row}'] = f'  {COGS_GROUP_LABELS_EN[sentinel]}'
                    ws[f'B{row}'] = ''
                    ws[f'C{row}'] = ''
                    for col_key in ('A', 'B', 'C'):
                        ws[f'{col_key}{row}'].font = cogs_sub_header_font
                        ws[f'{col_key}{row}'].fill = cogs_sub_header_fill
                        ws[f'{col_key}{row}'].border = thin_border

                    for i in range(12):
                        col_letter = get_column_letter(MONTH_START_COL + i)
                        cell = ws[f'{col_letter}{row}']
                        value = group.get('monthly', {}).get(i + 1, 0)
                        cell.value = value if value != 0 else None
                        cell.font = cogs_sub_header_font
                        cell.fill = cogs_sub_header_fill
                        cell.alignment = right_align
                        cell.number_format = '#,##0;(#,##0)'
                        cell.border = thin_border

                    annual_cell = ws[f'{annual_col}{row}']
                    annual_val = group.get('annual', 0)
                    annual_cell.value = annual_val if annual_val != 0 else None
                    annual_cell.font = cogs_sub_header_font
                    annual_cell.fill = cogs_sub_header_fill
                    annual_cell.alignment = right_align
                    annual_cell.number_format = '#,##0;(#,##0)'
                    annual_cell.border = thin_border
                    row += 1

                    for code in sorted(group.get('sub', {}).keys()):
                        leaf = group['sub'][code]
                        leaf_en = all_mappings.get(code, '')
                        ws[f'A{row}'] = f'      {leaf["name"]}'
                        ws[f'B{row}'] = code
                        ws[f'C{row}'] = leaf_en
                        for col_key in ('A', 'B', 'C'):
                            ws[f'{col_key}{row}'].font = cogs_leaf_font
                            ws[f'{col_key}{row}'].border = thin_border

                        for i in range(12):
                            col_letter = get_column_letter(MONTH_START_COL + i)
                            cell = ws[f'{col_letter}{row}']
                            value = leaf.get('monthly', {}).get(i + 1, 0)
                            cell.value = value if value != 0 else None
                            cell.font = cogs_leaf_font
                            cell.alignment = right_align
                            cell.number_format = '#,##0;(#,##0)'
                            cell.border = thin_border

                        annual_cell = ws[f'{annual_col}{row}']
                        leaf_annual = leaf.get('annual', 0)
                        annual_cell.value = leaf_annual if leaf_annual != 0 else None
                        annual_cell.font = cogs_leaf_font
                        annual_cell.alignment = right_align
                        annual_cell.number_format = '#,##0;(#,##0)'
                        annual_cell.border = thin_border
                        row += 1
            else:
                for code in sorted(subs.keys()):
                    sub = subs[code]
                    sub_en = all_mappings.get(code, '')
                    ws[f'A{row}'] = f'    {sub["name"]}'
                    ws[f'B{row}'] = code
                    ws[f'C{row}'] = sub_en
                    for col_key in ('A', 'B', 'C'):
                        ws[f'{col_key}{row}'].font = sub_font
                        ws[f'{col_key}{row}'].border = thin_border

                    for i in range(12):
                        col_letter = get_column_letter(MONTH_START_COL + i)
                        cell = ws[f'{col_letter}{row}']
                        value = sub.get('monthly', {}).get(i + 1, 0)
                        cell.value = value if value != 0 else None
                        cell.font = sub_font
                        cell.alignment = right_align
                        cell.number_format = '#,##0;(#,##0)'
                        cell.border = thin_border

                    annual_cell = ws[f'{annual_col}{row}']
                    annual_cell.value = sub.get('annual', 0) if sub.get('annual', 0) != 0 else None
                    annual_cell.font = sub_font
                    annual_cell.alignment = right_align
                    annual_cell.number_format = '#,##0;(#,##0)'
                    annual_cell.border = thin_border
                    row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 30
        for col in range(MONTH_START_COL, ANNUAL_COL + 1):
            ws.column_dimensions[get_column_letter(col)].width = 13

        # Return file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'P&L_Report_{year}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to export P&L report: {str(e)}'}), 500


@pl_api_bp.route('/delete-all', methods=['DELETE'])
@login_required
@app_access_required('pl')
def delete_all_pl_data():
    try:
        PLEntry.query.delete()
        PLSDCK.query.delete()
        db.session.commit()
        return jsonify({'message': 'All P&L data deleted'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to delete data: {str(e)}'}), 500


# ─── API: Account Mapping ────────────────────────────────────────────
@pl_api_bp.route('/mappings', methods=['GET'])
@login_required
@app_access_required('pl')
def get_mappings():
    """Get all account mappings."""
    try:
        mappings = AccountMapping.query.all()
        return jsonify({
            'mappings': [
                {'local_code': m.local_code, 'hq_code': m.hq_code}
                for m in mappings
            ]
        })
    except Exception as e:
        return jsonify({'error': f'Failed to fetch mappings: {str(e)}'}), 500


@pl_api_bp.route('/mappings/add', methods=['POST'])
@login_required
@app_access_required('pl')
def add_mapping():
    """Add a new account mapping."""
    data = request.get_json() or {}
    local_code = (data.get('local_code') or '').strip()
    hq_code = (data.get('hq_code') or '').strip()

    if not local_code or not hq_code:
        return jsonify({'error': 'Both local_code and hq_code are required'}), 400

    try:
        # Check if mapping already exists
        existing = AccountMapping.query.filter_by(local_code=local_code).first()
        if existing:
            return jsonify({'error': f'Mapping for {local_code} already exists'}), 400

        mapping = AccountMapping(local_code=local_code, hq_code=hq_code)
        db.session.add(mapping)
        db.session.commit()
        return jsonify({'message': f'Mapping added: {local_code} -> {hq_code}'}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to add mapping: {str(e)}'}), 500


@pl_api_bp.route('/mappings/<local_code>', methods=['DELETE'])
@login_required
@app_access_required('pl')
def delete_mapping(local_code):
    """Delete an account mapping."""
    try:
        mapping = AccountMapping.query.filter_by(local_code=local_code).first()
        if not mapping:
            return jsonify({'error': f'Mapping for {local_code} not found'}), 404

        db.session.delete(mapping)
        db.session.commit()
        return jsonify({'message': f'Mapping deleted: {local_code}'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to delete mapping: {local_code}: {str(e)}'}), 500


@pl_api_bp.route('/mappings/import-excel', methods=['POST'])
@login_required
@app_access_required('pl')
def import_mappings_excel():
    """Import account mappings from Excel file."""
    import pandas as pd

    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'Only .xlsx and .xls files are allowed'}), 400

    filename = secure_filename(file.filename)
    upload_folder = current_app.config['UPLOAD_FOLDER']
    os.makedirs(upload_folder, exist_ok=True)
    filepath = os.path.join(upload_folder, filename)
    file.save(filepath)

    try:
        # Detect file type
        is_xls = filepath.lower().endswith('.xls')
        engine = 'xlrd' if is_xls else 'openpyxl'

        # Read Excel file
        df = pd.read_excel(filepath, engine=engine)
        df.columns = df.columns.str.strip()

        # Find columns (try common names)
        local_col = None
        hq_col = None

        for col in df.columns:
            col_lower = str(col).lower().strip()
            if 'local' in col_lower or 'local code' in col_lower:
                local_col = col
            if 'hq' in col_lower or 'map' in col_lower or 'hq code' in col_lower:
                hq_col = col

        # If not found, use first two columns
        if not local_col or not hq_col:
            cols = list(df.columns)
            if len(cols) >= 2:
                local_col = cols[0]
                hq_col = cols[1]
            else:
                return jsonify({'error': 'File must have at least 2 columns'}), 400

        success_count = 0
        error_count = 0
        errors = []

        for idx, row in df.iterrows():
            try:
                # Fix 1: strip .0 from numeric codes (pandas reads as float)
                raw = str(row.get(local_col, '')).strip()
                try:
                    local_code = str(int(float(raw)))
                except (ValueError, TypeError):
                    local_code = raw

                hq_code = str(row.get(hq_col, '')).strip()

                if not local_code or not hq_code or local_code == 'nan':
                    continue

                # Fix 2: upsert instead of skip
                existing = AccountMapping.query.filter_by(local_code=local_code).first()
                if existing:
                    existing.hq_code = hq_code
                else:
                    db.session.add(AccountMapping(local_code=local_code, hq_code=hq_code))
                success_count += 1
            except Exception as e:
                error_count += 1
                errors.append(f"Row {idx + 2}: {str(e)}")

        db.session.commit()

        return jsonify({
            'message': f'Imported {success_count} mappings, {error_count} errors',
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors[:20],
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Import failed: {str(e)}'}), 500
    finally:
        if os.path.exists(filepath):
            os.remove(filepath)


@pl_api_bp.route('/mappings/import-google', methods=['POST'])
@login_required
@app_access_required('pl')
def import_mappings_google():
    """Import account mappings from Google Sheets."""
    import pandas as pd
    import tempfile
    import urllib.request

    data = request.get_json() or {}
    url = (data.get('url') or '').strip()

    if not url:
        return jsonify({'error': 'No Google Sheets URL provided'}), 400

    if 'docs.google.com/spreadsheets' not in url:
        return jsonify({'error': 'Invalid Google Sheets URL'}), 400

    try:
        # Extract spreadsheet ID
        import re
        match = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', url)
        if not match:
            return jsonify({'error': 'Invalid Google Sheets URL'}), 400

        sheet_id = match.group(1)
        export_url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx'

        # Download file
        tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')

        try:
            urllib.request.urlretrieve(export_url, tmp_path)

            # Read Excel file
            df = pd.read_excel(tmp_path, engine='openpyxl')
            df.columns = df.columns.str.strip()

            # Find columns
            local_col = None
            hq_col = None

            for col in df.columns:
                col_lower = str(col).lower().strip()
                if 'local' in col_lower or 'local code' in col_lower:
                    local_col = col
                if 'hq' in col_lower or 'map' in col_lower or 'hq code' in col_lower:
                    hq_col = col

            # If not found, use first two columns
            if not local_col or not hq_col:
                cols = list(df.columns)
                if len(cols) >= 2:
                    local_col = cols[0]
                    hq_col = cols[1]
                else:
                    return jsonify({'error': 'Sheet must have at least 2 columns'}), 400

            success_count = 0
            error_count = 0
            errors = []

            for idx, row in df.iterrows():
                try:
                    # Fix 1: strip .0 from numeric codes (pandas reads as float)
                    raw = str(row.get(local_col, '')).strip()
                    try:
                        local_code = str(int(float(raw)))
                    except (ValueError, TypeError):
                        local_code = raw

                    hq_code = str(row.get(hq_col, '')).strip()

                    if not local_code or not hq_code or local_code == 'nan':
                        continue

                    # Fix 2: upsert instead of skip
                    existing = AccountMapping.query.filter_by(local_code=local_code).first()
                    if existing:
                        existing.hq_code = hq_code
                    else:
                        db.session.add(AccountMapping(local_code=local_code, hq_code=hq_code))
                    success_count += 1
                except Exception as e:
                    error_count += 1
                    errors.append(f"Row {idx + 2}: {str(e)}")

            db.session.commit()

            return jsonify({
                'message': f'Imported {success_count} mappings, {error_count} errors',
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors[:20],
            })
        finally:
            os.close(tmp_fd)
            os.unlink(tmp_path)
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Import failed: {str(e)}'}), 500


@pl_api_bp.route('/mappings/export', methods=['GET'])
@login_required
@app_access_required('pl')
def export_mappings():
    """Export account mappings to Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    try:
        mappings = AccountMapping.query.all()

        if not mappings:
            return jsonify({'error': 'No mappings to export'}), 400

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Mappings'

        # Define styles
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # Header row
        headers = ['Local Code', 'HQ Code']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Data rows
        for row_idx, mapping in enumerate(mappings, 2):
            ws.cell(row=row_idx, column=1).value = mapping.local_code
            ws.cell(row=row_idx, column=2).value = mapping.hq_code

            for col_idx in range(1, 3):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25

        # Return file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Account_Mappings.xlsx',
        )
    except Exception as e:
        return jsonify({'error': f'Export failed: {str(e)}'}), 500
