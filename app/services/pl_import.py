import os
import re
import tempfile
import urllib.request
import pandas as pd
from datetime import datetime
from app import db
from app.models.pl_entry import PLEntry
from app.models.pl_sdck import PLSDCK

COLUMN_MAP = {
    'Tk đ.ứng': 'account_code',
    'Tên tk đ.ứng': 'account_name',
    'Ps nợ': 'debit',
    'Ps có': 'credit',
}

# Parent account codes to exclude (3-digit codes)
PARENT_CODES = {'511', '515', '642', '635', '711', '811', '154'}


def detect_file_type(filename):
    """Detect file type from filename: '911', '154', or 'SDCK'."""
    if 'SDCK' in filename.upper():
        # Extract account number from SDCK filename: "SDCK 1541 01.02.2026"
        match = re.search(r'SDCK\s+(\d+)', filename, re.IGNORECASE)
        if match:
            return ('SDCK', match.group(1))
        else:
            raise ValueError(f"Cannot extract account number from SDCK filename: '{filename}'")
    elif '911' in filename:
        return '911'
    elif '154' in filename:
        return '154'
    else:
        raise ValueError(
            f"Cannot detect file type from filename: '{filename}'\n"
            f"Filename must contain '911', '154', or 'SDCK 1541' and date pattern.\n"
            f"Expected format: '911 t 01.02.2026-28.02.2026' or 'SDCK 1541 01.02.2026-28.02.2026'"
        )


def extract_month_year(filename):
    """
    Extract month and year from filename.
    Expects format like: "911 t 01.02.2026-28.02.2026" or "154 t 01-02-2026-28-02-2026"
    Returns (month, year) tuple.
    """
    # Regex for date patterns: DD.MM.YYYY or DD-MM-YYYY (with optional time)
    match = re.search(r'(\d{2})[.\-](\d{2})[.\-](\d{4})', filename)
    if not match:
        raise ValueError(
            f"Cannot extract month/year from filename: '{filename}'\n"
            f"Expected format: '911 t 01.02.2026-28.02.2026' or '154 t 01-02-2026-28-02-2026'"
        )

    day_str, month_str, year_str = match.groups()
    month = int(month_str)
    year = int(year_str)

    if not (1 <= month <= 12):
        raise ValueError(f"Invalid month {month} extracted from filename: {filename}")

    return month, year


def find_header_row(filepath):
    """
    Find the row number containing the header.
    Scans first 10 rows looking for 'Tk đ.ứng' or 'Ps nợ'.
    Returns (header_row_index, skiprows_count) or raises ValueError.
    """
    try:
        # Detect file type and choose appropriate engine
        is_xls = filepath.lower().endswith('.xls')
        engine = 'xlrd' if is_xls else 'openpyxl'

        # Read first 10 rows without header to scan
        df_scan = pd.read_excel(filepath, nrows=10, header=None, engine=engine)

        for idx, row in df_scan.iterrows():
            # Check if any cell contains expected column names
            row_str = ' '.join([str(cell).strip() for cell in row if pd.notna(cell)])
            if 'Tk đ.ứng' in row_str or 'Ps nợ' in row_str or 'Ps no' in row_str:
                return idx, idx  # (header_row_index, skiprows=idx)

        # Fallback: check if first 3 rows are skipped (like OTL)
        df_check = pd.read_excel(filepath, skiprows=3, nrows=1, engine=engine)
        found_cols = [col for col in COLUMN_MAP.keys() if col in df_check.columns or _find_column(df_check, col)]
        if found_cols:
            return 3, 3

        # Fallback 2: Try row 0 as header
        df_check2 = pd.read_excel(filepath, nrows=1, engine=engine)
        found_cols2 = [col for col in COLUMN_MAP.keys() if col in df_check2.columns or _find_column(df_check2, col)]
        if found_cols2:
            return 0, 0

        raise ValueError("Cannot find header row in Excel file. Expected columns: Tk đ.ứng, Tên tk đ.ứng, Ps nợ, Ps có")
    except Exception as e:
        raise ValueError(f"Error scanning for header row: {str(e)}")


def get_leaf_codes(codes):
    """
    From a list of account codes, return only the leaf codes (codes with no children).

    A code is a leaf if no other code in the set starts with it.
    Example: given {511, 5113, 51131, 511311, 511312}, leaves are {511311, 511312}
    """
    codes_set = set()
    for c in codes:
        if c and not pd.isna(c):
            codes_set.add(str(c).strip())

    codes_set.discard('')  # Remove empty strings

    if not codes_set:
        return set()

    leaves = set()
    for code in codes_set:
        # Check if any OTHER code starts with this one (would be a child)
        is_leaf = not any(
            other != code and str(other).startswith(code)
            for other in codes_set
        )
        if is_leaf:
            leaves.add(code)

    return leaves


def parse_numeric(value):
    """Parse numeric value, return 0 for invalid."""
    if pd.isna(value) or value is None:
        return 0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0


def _find_column(df, column_name):
    """Find column by name, handling variations like extra spaces or unicode."""
    # Try exact match first
    if column_name in df.columns:
        return column_name

    # Try case-insensitive and whitespace-normalized match
    normalized_target = column_name.strip().lower()
    for col in df.columns:
        # Skip NaN or non-string column names
        if pd.isna(col):
            continue
        col_str = str(col).strip()
        if col_str.lower() == normalized_target:
            return col

    return None


def _import_dataframe(df, file_type, month, year, source_name):
    """Import a DataFrame into PLEntry. Returns (success, error_count, errors)."""
    # Strip whitespace from column names
    df.columns = df.columns.str.strip()

    # Delete existing entries for this month/year/file_type
    PLEntry.query.filter_by(month=month, year=year, file_type=file_type).delete()

    success_count = 0
    error_count = 0
    skipped_count = 0
    errors = []

    # Find actual column names in the dataframe
    col_tk = _find_column(df, 'Tk đ.ứng')
    col_ten = _find_column(df, 'Tên tk đ.ứng')
    col_no = _find_column(df, 'Ps nợ')
    col_co = _find_column(df, 'Ps có')

    # Debug: check columns
    if not all([col_tk, col_ten, col_no, col_co]):
        missing = []
        if not col_tk: missing.append('Tk đ.ứng')
        if not col_ten: missing.append('Tên tk đ.ứng')
        if not col_no: missing.append('Ps nợ')
        if not col_co: missing.append('Ps có')
        errors.append(f"ERROR: Missing columns: {', '.join(missing)}. Available: {list(df.columns)}")
        return 0, len(errors), errors

    for idx, row in df.iterrows():
        try:
            # Excel reads numeric-looking codes as floats (e.g. 51131 → 51131.0)
            # Convert to int first to strip .0 suffix, enabling correct leaf detection
            account_code_raw = row.get(col_tk, '')
            if pd.notna(account_code_raw):
                code_str = str(account_code_raw).strip()
                try:
                    # int(51131.0) = 51131 → str(51131) = "51131" (clean)
                    account_code = str(int(float(code_str)))
                except (ValueError, TypeError):
                    # Non-numeric codes (e.g. "5113A") — keep as-is
                    account_code = code_str
            else:
                account_code = ''

            account_name = str(row.get(col_ten, '')).strip()
            debit = parse_numeric(row.get(col_no))
            credit = parse_numeric(row.get(col_co))

            # Skip empty rows (no account code or no debit/credit)
            if not account_code or (debit == 0 and credit == 0):
                skipped_count += 1
                continue

            entry = PLEntry(
                source_file=source_name,
                file_type=file_type,
                month=month,
                year=year,
                account_code=account_code,
                account_name=account_name,
                debit=debit,
                credit=credit,
                imported_at=datetime.utcnow(),
            )

            db.session.add(entry)
            success_count += 1

        except Exception as e:
            error_count += 1
            errors.append(f"Row {idx + 1}: {str(e)}")

    db.session.commit()

    # Add info about skipped rows
    if skipped_count > 0:
        errors.append(f"INFO: {skipped_count} rows skipped (parent accounts or empty)")

    return success_count, error_count, errors[:20]  # Return first 20 errors


def import_pl_file(file_path, filename):
    """
    Import P&L data from an Excel file.
    Detects file type, extracts month/year, finds header row, and imports.
    Returns (success_count, error_count, errors).
    """
    file_type_result = detect_file_type(filename)

    # Handle SDCK detection (returns tuple)
    if isinstance(file_type_result, tuple):
        file_type, account_target = file_type_result
        return import_sdck_file(file_path, filename, account_target)
    else:
        file_type = file_type_result

    month, year = extract_month_year(filename)
    header_idx, skiprows = find_header_row(file_path)

    # Choose engine based on file extension
    is_xls = file_path.lower().endswith('.xls')
    engine = 'xlrd' if is_xls else 'openpyxl'

    df = pd.read_excel(file_path, skiprows=skiprows, engine=engine)

    return _import_dataframe(df, file_type, month, year, filename)


def import_sdck_file(file_path, filename, account_target):
    """
    Import SDCK (Số dư cuối kỳ / Ending Balance) data from an Excel file.
    Format: Filename like "SDCK 1541 01.02.2026-28.02.2026"
    Columns: "Tk", "Tên tk", "Ps nợ"

    Returns (success_count, error_count, errors).
    """
    month, year = extract_month_year(filename)
    header_idx, skiprows = find_header_row(file_path)

    # Choose engine based on file extension
    is_xls = file_path.lower().endswith('.xls')
    engine = 'xlrd' if is_xls else 'openpyxl'

    df = pd.read_excel(file_path, skiprows=skiprows, engine=engine)

    success_count = 0
    error_count = 0
    errors = []

    # Find columns
    col_tk = None
    col_ten = None
    col_no = None

    for col in df.columns:
        col_str = str(col).strip().lower()
        if 'tk' in col_str and col_tk is None:
            col_tk = col
        if 'tên' in col_str or 'ten' in col_str:
            col_ten = col
        if ('nợ' in col_str or 'no' in col_str) and col_no is None:
            col_no = col

    if not col_tk or not col_no:
        errors.append(f"ERROR: Missing required columns. Expected: 'Tk', 'Tên tk', 'Ps nợ'")
        return 0, len(errors), errors

    for idx, row in df.iterrows():
        try:
            # Extract account code
            account_code_raw = row.get(col_tk, '')
            if pd.notna(account_code_raw):
                code_str = str(account_code_raw).strip()
                try:
                    account_code = str(int(float(code_str)))
                except (ValueError, TypeError):
                    account_code = code_str
            else:
                account_code = ''

            account_name = str(row.get(col_ten, '')).strip() if col_ten else ''
            balance_raw = row.get(col_no, 0)

            # Parse balance as float (store as POSITIVE - will be subtracted in COGS formula)
            try:
                balance = float(balance_raw) if pd.notna(balance_raw) else 0
            except (ValueError, TypeError):
                balance = 0

            # Skip empty rows
            if not account_code or balance == 0:
                continue

            # Create SDCK entry
            entry = PLSDCK(
                source_file=filename,
                account_target=account_target,
                month=month,
                year=year,
                account_code=account_code,
                account_name=account_name,
                balance=balance,
                imported_at=datetime.utcnow(),
            )

            db.session.add(entry)
            success_count += 1

        except Exception as e:
            error_count += 1
            errors.append(f"Row {idx + 1}: {str(e)}")

    db.session.commit()

    if success_count > 0:
        errors.append(f"INFO: Imported {success_count} SDCK entries for account {account_target}")

    return success_count, error_count, errors[:20]


def _import_sdck_dataframe(df, account_target, month, year, source_name):
    """
    Import SDCK (Số dư cuối kỳ / Ending Balance) data from a DataFrame.
    Columns: "Tk", "Tên tk", "Ps nợ"

    Returns (success_count, error_count, errors).
    """
    success_count = 0
    error_count = 0
    errors = []

    # Find columns
    col_tk = None
    col_ten = None
    col_no = None

    for col in df.columns:
        col_str = str(col).strip().lower()
        if 'tk' in col_str and col_tk is None:
            col_tk = col
        if 'tên' in col_str or 'ten' in col_str:
            col_ten = col
        if ('nợ' in col_str or 'no' in col_str) and col_no is None:
            col_no = col

    if not col_tk or not col_no:
        errors.append(f"ERROR: Missing required columns. Expected: 'Tk', 'Tên tk', 'Ps nợ'")
        return 0, len(errors), errors

    for idx, row in df.iterrows():
        try:
            # Extract account code
            account_code_raw = row.get(col_tk, '')
            if pd.notna(account_code_raw):
                code_str = str(account_code_raw).strip()
                try:
                    account_code = str(int(float(code_str)))
                except (ValueError, TypeError):
                    account_code = code_str
            else:
                account_code = ''

            account_name = str(row.get(col_ten, '')).strip() if col_ten else ''
            balance_raw = row.get(col_no, 0)

            # Parse balance as float (store as POSITIVE - will be subtracted in COGS formula)
            try:
                balance = float(balance_raw) if pd.notna(balance_raw) else 0
            except (ValueError, TypeError):
                balance = 0

            # Skip empty rows
            if not account_code or balance == 0:
                continue

            # Create SDCK entry
            entry = PLSDCK(
                source_file=source_name,
                account_target=account_target,
                month=month,
                year=year,
                account_code=account_code,
                account_name=account_name,
                balance=balance,
                imported_at=datetime.utcnow(),
            )

            db.session.add(entry)
            success_count += 1

        except Exception as e:
            error_count += 1
            errors.append(f"Row {idx + 1}: {str(e)}")

    db.session.commit()

    if success_count > 0:
        errors.append(f"INFO: Imported {success_count} SDCK entries for account {account_target}")

    return success_count, error_count, errors[:20]


def import_pl_google_sheet(url, month=None, year=None):
    """
    Import P&L data from a public Google Sheets link.
    Handles URLs with or without query parameters.

    Args:
        url: Google Sheets URL
        month: (optional) Override month detection (1-12)
        year: (optional) Override year detection

    Returns (success_count, error_count, errors).
    """
    # Extract spreadsheet ID - handles URLs with /edit?gid=... or #gid=...
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', url)
    if not match:
        raise ValueError('Invalid Google Sheets URL. Expected format: https://docs.google.com/spreadsheets/d/SPREADSHEET_ID/...')

    sheet_id = match.group(1)

    # Download as xlsx
    export_url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx'
    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')

    try:
        urllib.request.urlretrieve(export_url, tmp_path)

        # Try to get sheet title from API first
        sheet_title = None
        try:
            api_url = f'https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}?fields=properties.title&key=AIzaSyBZL-5k-M47cPVU1i6Vc-NP0-yMdlKHQP4'
            response = urllib.request.urlopen(api_url, timeout=5)
            import json
            data = json.loads(response.read())
            sheet_title = data.get('properties', {}).get('title', '')
        except Exception:
            pass

        # If API didn't work, try to extract sheet name from the downloaded Excel file
        if not sheet_title:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(tmp_path)
                # Get the first sheet's name (Google Sheets exports first sheet)
                if wb.sheetnames:
                    sheet_title = wb.sheetnames[0]
            except Exception:
                pass

        # Auto-detect file type from sheet title first (like filename), then from content
        file_type = None
        account_target = None  # For SDCK files
        if sheet_title:
            try:
                file_type_result = detect_file_type(sheet_title)
                # Handle SDCK detection (returns tuple)
                if isinstance(file_type_result, tuple):
                    file_type, account_target = file_type_result
                else:
                    file_type = file_type_result
            except ValueError:
                pass  # Sheet title doesn't contain file type marker

        # Fallback to content detection if title didn't work
        if not file_type:
            file_type = detect_file_type_from_content(tmp_path)

        # Try to extract month/year from sheet title first (e.g., "911 t 01.02.2026-28.02.2026")
        extracted_month = None
        extracted_year = None
        if sheet_title:
            title_match = re.search(r'(\d{2})[.\-](\d{2})[.\-](\d{4})', sheet_title)
            if title_match:
                extracted_month = int(title_match.group(2))
                extracted_year = int(title_match.group(3))

        # Use provided month/year, then try sheet title, then content
        if month is None or year is None:
            if extracted_month and extracted_year:
                month = month or extracted_month
                year = year or extracted_year
            else:
                detected_month, detected_year = extract_month_year_from_content(tmp_path)
                if detected_month and detected_year:
                    month = month or detected_month
                    year = year or detected_year
                else:
                    # Error if we couldn't extract month/year from any source
                    raise ValueError(
                        f"Cannot extract month/year. Please ensure:\n"
                        f"1. Sheet name contains date pattern (e.g., '911 t 01.02.2026-28.02.2026')\n"
                        f"2. Or provide month/year manually in the import form"
                    )

        header_idx, skiprows = find_header_row(tmp_path)

        df = pd.read_excel(tmp_path, skiprows=skiprows, engine='openpyxl')
        source_name = f'google-sheet:{sheet_id}'

        # Handle SDCK import separately
        if file_type == 'SDCK' and account_target:
            return _import_sdck_dataframe(df, account_target, month, year, source_name)
        else:
            return _import_dataframe(df, file_type, month, year, source_name)
    finally:
        os.close(tmp_fd)
        os.unlink(tmp_path)


def detect_file_type_from_content(filepath):
    """
    Detect file type from cell content.
    Scans first 20 rows for '911' or '154' text.
    Falls back to checking account codes if file type markers not found.
    """
    try:
        is_xls = filepath.lower().endswith('.xls')
        engine = 'xlrd' if is_xls else 'openpyxl'
        df_scan = pd.read_excel(filepath, nrows=20, header=None, engine=engine)

        # Check first 20 rows for file type markers
        for row in df_scan.values:
            row_str = ' '.join([str(cell) for cell in row])
            if '911' in row_str:
                return '911'
            elif '154' in row_str:
                return '154'

        # Fallback: check account codes in 4th column area
        for row in df_scan.iterrows():
            for cell in row[1]:
                cell_str = str(cell).strip()
                # Check if cell contains account code starting patterns
                if cell_str.startswith('154'):
                    return '154'
                elif cell_str.startswith('911'):
                    return '911'
                elif cell_str.startswith(('511', '515', '642', '635', '711', '811')):
                    return '911'  # Assume 911 if we find standard P&L account codes

        raise ValueError("Cannot detect file type from content - no '911' or '154' markers found")
    except Exception as e:
        raise ValueError(f"Error detecting file type: {str(e)}")


def extract_month_year_from_content(filepath):
    """
    Extract month/year from cell content.
    Scans first 20 rows for date pattern.
    Returns (month, year) or (None, None) if not found.
    """
    try:
        df_scan = pd.read_excel(filepath, nrows=20, header=None, engine='openpyxl')

        for row in df_scan.values:
            row_str = ' '.join([str(cell).strip() for cell in row])
            match = re.search(r'(\d{2})[.\-](\d{2})[.\-](\d{4})', row_str)
            if match:
                month = int(match.group(2))
                year = int(match.group(3))
                if 1 <= month <= 12:
                    return month, year

        # Return None if not found (will use default)
        return None, None
    except Exception as e:
        # Return None on error (will use default)
        return None, None


def calculate_pl(year):
    """
    Calculate P&L report for a given year, including sub-account breakdown.
    Returns dict: {month: {'totals': {...}, 'sub_accounts': {...}}, ...}

    Aggregation rules:
    - file_type='911': Revenue (511x credit, 515x credit, 711x credit) + Expenses (642x debit, 635x debit, 811x debit)
    - file_type='154' (Cost of Production): Giá vốn hàng bán from accounts 6xx (debit side)
    - Leaf-node filtering: only include codes that have no children within their (file, prefix) group
    - SDCK 1541: Ending inventory adjustment (subtracted from COGS)

    Account mapping:
    - Giá vốn hàng bán (COGS): TK 154 t - TK 6xx (Ps Nợ / Debit) + SDCK 1541 Inventory
    """
    entries = PLEntry.query.filter_by(year=year).all()

    # Map account prefix to P&L line item
    PREFIX_TO_LINE = {
        '511': 'Doanh thu thuần',
        '515': 'Doanh thu HĐTC',
        '711': 'Thu nhập khác',
        '642': 'Chi phí QLDN',
        '635': 'Chi phí tài chính',
        '811': 'Chi phí khác',
    }

    # Group by month
    monthly_data = {}
    for month in range(1, 13):
        monthly_data[month] = {
            'totals': {
                'Doanh thu thuần': 0,
                'Giá vốn hàng bán': 0,
                'Doanh thu HĐTC': 0,
                'Chi phí QLDN': 0,
                'Chi phí tài chính': 0,
                'Thu nhập khác': 0,
                'Chi phí khác': 0,
            },
            'sub_accounts': {
                'Doanh thu thuần': {},
                'Giá vốn hàng bán': {},
                'Doanh thu HĐTC': {},
                'Chi phí QLDN': {},
                'Chi phí tài chính': {},
                'Thu nhập khác': {},
                'Chi phí khác': {},
            }
        }

    # Pre-calculate leaf codes per (month, file_type, prefix group)
    leaf_cache = {}

    def get_leaf_set(month, file_type, prefix):
        """Get leaf codes for a specific (month, file_type, prefix) group."""
        key = (month, file_type, prefix)
        if key not in leaf_cache:
            # Query all codes for this group
            query = PLEntry.query.filter_by(year=year, month=month, file_type=file_type)
            codes_for_group = set()

            for entry in query.all():
                code = str(entry.account_code).strip()
                # Check if code matches this prefix
                if prefix == 'COGS_6':
                    # File 154 t: Only 6xx accounts (not 8xx)
                    if code[0] == '6':
                        codes_for_group.add(code)
                else:
                    if code.startswith(prefix):
                        codes_for_group.add(code)

            leaf_cache[key] = get_leaf_codes(codes_for_group)

        return leaf_cache[key]

    # Aggregate by account code
    for entry in entries:
        if entry.month not in monthly_data:
            continue

        month_data = monthly_data[entry.month]
        code = str(entry.account_code).strip()

        # Determine which line item and group this entry belongs to
        line_item = None
        is_leaf = False
        value = 0

        if entry.file_type == '911':
            # File 911: revenues + expenses from main accounts
            for prefix, item in PREFIX_TO_LINE.items():
                if code.startswith(prefix):
                    line_item = item
                    leaf_set = get_leaf_set(entry.month, '911', prefix)
                    is_leaf = code in leaf_set
                    # Determine value: credit for revenue, debit for expense
                    if prefix in ('511', '515', '711'):
                        value = entry.credit
                    else:
                        value = entry.debit
                    break

        elif entry.file_type == '154':
            # File "154 t" (Cost of Production): Giá vốn hàng bán from accounts 6xx (debit)
            # Note: SDCK 1541 ending inventory adjustment is added separately below
            if code[0] == '6':
                line_item = 'Giá vốn hàng bán'
                leaf_set = get_leaf_set(entry.month, '154', 'COGS_6')
                is_leaf = code in leaf_set
                value = entry.debit

        if line_item and is_leaf and value != 0:
            # Add to totals
            month_data['totals'][line_item] += value

            # Add to sub_accounts
            if code not in month_data['sub_accounts'][line_item]:
                month_data['sub_accounts'][line_item][code] = {
                    'name': entry.account_name or '',
                    'value': 0
                }
            month_data['sub_accounts'][line_item][code]['value'] += value

    # Process SDCK (Số dư cuối kỳ / Ending Balance) data - add to COGS
    sdck_entries = PLSDCK.query.filter_by(year=year).all()
    for entry in sdck_entries:
        if entry.month in monthly_data:
            month_data = monthly_data[entry.month]
            # SDCK data represents ending balance of accounts, add to COGS
            line_item = 'Giá vốn hàng bán'
            code = str(entry.account_code).strip()
            value = entry.balance  # Already negative if needed

            if value != 0:
                month_data['totals'][line_item] += value

                if code not in month_data['sub_accounts'][line_item]:
                    month_data['sub_accounts'][line_item][code] = {
                        'name': entry.account_name or '',
                        'value': 0
                    }
                month_data['sub_accounts'][line_item][code]['value'] += value

    # Calculate P&L metrics for each month
    result = {}
    for month, data in monthly_data.items():
        result[month] = calculate_monthly_pl(
            data['totals'],
            data['sub_accounts']
        )

    return result


def calculate_monthly_pl(totals, sub_accounts):
    """
    Calculate P&L line items from aggregated monthly totals and sub-accounts.
    Returns: {
        'Doanh thu thuần': value,
        'Giá vốn hàng bán': value,
        ...
        'Lợi nhuận trước thuế': value,
        '__sub_accounts__': {
            'Doanh thu thuần': {code: {name, value}, ...},
            ...
        }
    }
    """
    revenue = totals['Doanh thu thuần']
    cogs = totals['Giá vốn hàng bán']
    sga = totals['Chi phí QLDN']
    fin_revenue = totals['Doanh thu HĐTC']
    fin_expense = totals['Chi phí tài chính']
    other_revenue = totals['Thu nhập khác']
    other_expense = totals['Chi phí khác']

    gross_profit = revenue - cogs
    operating_profit = gross_profit - sga
    pbt = operating_profit + (fin_revenue - fin_expense) + (other_revenue - other_expense)

    result = {
        'Doanh thu thuần': revenue,
        'Giá vốn hàng bán': cogs,
        'Lợi nhuận gộp': gross_profit,
        'Chi phí QLDN': sga,
        'Lợi nhuận thuần từ HĐKD': operating_profit,
        'Doanh thu HĐTC': fin_revenue,
        'Chi phí tài chính': fin_expense,
        'Thu nhập khác': other_revenue,
        'Chi phí khác': other_expense,
        'Lợi nhuận trước thuế': pbt,
        '__sub_accounts__': sub_accounts,
    }
    return result


def get_available_years():
    """Get list of years that have P&L data."""
    years = db.session.query(PLEntry.year).distinct().all()
    return sorted([y[0] for y in years if y[0] is not None])
